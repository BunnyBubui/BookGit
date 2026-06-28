import os
import glob
import pandas as pd
from config.Config import folder_path, target_locations, save_dir, Colors

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print(f"{Colors.RED}❌ ไม่พบไลบรารี openpyxl กรุณาพิมพ์คำสั่ง: pip install openpyxl{Colors.RESET}")
    exit()

def DiagnoseDataHealth():
    os.system('cls' if os.name == 'nt' else 'clear')
    print(f"{Colors.BOLD}{Colors.CYAN}================================================================={Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN} 🩺 MASTER DATA HEALTH DIAGNOSTIC & EXCEL AUDIT GENERATOR{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}=================================================================\n{Colors.RESET}")

    all_files = glob.glob(os.path.join(folder_path, "*.csv"))
    if not all_files:
        print(f"{Colors.RED}❌ ไม่พบไฟล์ CSV ในโฟลเดอร์: {folder_path}{Colors.RESET}"); return

    print(f"{Colors.CYAN}⏳ [1/4] Scanning & Grouping 26M+ raw timestamps into Daily Hash Maps...{Colors.RESET}")
    
    # ดึงข้อมูลดิบเข้ามาเก็บแยกลง Dictionary ความเร็วสูง O(1)
    raw_streams = {}
    for loc in target_locations:
        matched = [f for f in all_files if loc in os.path.basename(f)]
        df_list = []
        for f in matched:
            try:
                t = pd.read_csv(f, on_bad_lines='skip')
                if 'datetime' in t.columns:
                    t['dt'] = pd.to_datetime(t['datetime'], format='%d-%m-%Y-%H-%M-%S', errors='coerce')
                    if t['dt'].isna().all(): t['dt'] = pd.to_datetime(t['datetime'], errors='coerce')
                    df_list.append(t.dropna(subset=['dt']))
            except: pass
            
        if df_list:
            combined = pd.concat(df_list).sort_values('dt')
            combined['date_only'] = combined['dt'].dt.floor('D')
            # แปลง Groupby เป็น Dict เพื่อสปีดตอนสแกนทีละวัน
            raw_streams[loc] = dict(tuple(combined.groupby('date_only')))

    if not raw_streams:
        print(f"{Colors.RED}❌ ไม่สามารถอ่านข้อมูลวันที่จากไฟล์ใดๆ ได้เลย{Colors.RESET}"); return

    # หาขอบเขตวันที่ทั้งหมดตั้งแต่ต้นจนจบโปรเจกต์
    all_dates = []
    for stream in raw_streams.values(): all_dates.extend(list(stream.keys()))
    min_date, max_date = min(all_dates), max(all_dates)
    master_calendar = pd.date_range(start=min_date, end=max_date, freq='D')

    print(f"{Colors.CYAN}⏳ [2/4] Analyzing Gaps, Duplicates & Overlaps day-by-day ({len(master_calendar)} days)...{Colors.RESET}")

    matrix_rows = []
    summary_stats = {loc: {'healthy':0, 'gaps':0, 'excess':0, 'missing':0, 'active_days':0} for loc in target_locations}

    for curr_day in master_calendar:
        row_text = {'Date': curr_day.strftime('%Y-%m-%d (%a)')}
        row_status = {}

        # เช็คว่าวันนี้ตึกเปิดทำงานหรือไม่ (มีเซนเซอร์ตัวไหนส่งค่ามาบ้างไหม)
        is_building_active_today = any(curr_day in raw_streams.get(loc, {}) for loc in target_locations)

        for loc in target_locations:
            df_day = raw_streams.get(loc, {}).get(curr_day, None)
            loc_key = loc.upper()

            if df_day is None or len(df_day) == 0:
                if is_building_active_today:
                    row_text[loc_key] = "- (ไม่มีข้อมูลวันนี้)"
                    row_status[loc_key] = "MISSING"
                    summary_stats[loc]['missing'] += 1
                else:
                    row_text[loc_key] = "⚫ ปิดตึก/ระบบหยุดทำงานทั้งหมด"
                    row_status[loc_key] = "BLACKOUT"
            else:
                summary_stats[loc]['active_days'] += 1
                df_sorted = df_day.sort_values('dt')
                
                # 1. เช็คแถวซ้ำ (Duplicate rows)
                dups = df_sorted['dt'].duplicated().sum()

                # 2. เช็คขาดช่วง > 15 นาที
                diffs = df_sorted['dt'].diff().dt.total_seconds() / 60.0
                gaps = diffs[diffs > 15.0]

                if len(gaps) > 0:
                    w_idx = gaps.idxmax()
                    t1 = df_sorted['dt'].loc[w_idx - 1].strftime('%H:%M')
                    t2 = df_sorted['dt'].loc[w_idx].strftime('%H:%M')
                    row_text[loc_key] = f"🟡 ขาด {len(gaps)} จุด (หนักสุด {t1}-{t2} หายไป {gaps.max():.0f}น.)"
                    row_status[loc_key] = "GAPS"
                    summary_stats[loc]['gaps'] += 1
                elif dups > 0:
                    row_text[loc_key] = f"🟣 ข้อมูลเกิน/ซ้ำ (+{dups} แถวซ้ำหลังเที่ยงคืน)"
                    row_status[loc_key] = "EXCESS"
                    summary_stats[loc]['excess'] += 1
                else:
                    row_text[loc_key] = f"🟢 สมบูรณ์ ({len(df_day):,} แถว)"
                    row_status[loc_key] = "HEALTHY"
                    summary_stats[loc]['healthy'] += 1

        matrix_rows.append((row_text, row_status))

    print(f"{Colors.CYAN}⏳ [3/4] Constructing OpenPyXL Workbook & Injecting Design Styles...{Colors.RESET}")

    excel_path = os.path.join(save_dir, "Data_Health_Master_Audit.xlsx")
    wb = openpyxl.Workbook()

    # =========================================================================
    # 🎨 STYLING DEFINITIONS
    # =========================================================================
    thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                         top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

    fills = {
        "HEALTHY": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "GAPS":    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "EXCESS":  PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),
        "MISSING": PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid"),
        "BLACKOUT":PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    }
    fonts = {
        "HEALTHY": Font(name="Segoe UI", size=10, color="375623", bold=True),
        "GAPS":    Font(name="Segoe UI", size=10, color="B25900", bold=True),
        "EXCESS":  Font(name="Segoe UI", size=10, color="371435", bold=True),
        "MISSING": Font(name="Segoe UI", size=10, color="9C0006", bold=True),
        "BLACKOUT":Font(name="Segoe UI", size=10, color="7F7F7F", italic=True)
    }

    # =========================================================================
    # 📊 SHEET 1: EXECUTIVE SUMMARY
    # =========================================================================
    ws_sum = wb.active
    ws_sum.title = "📊 Executive Summary"
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum.append(["AI DATASET HEALTH EXECUTIVE AUDIT DASHBOARD"])
    ws_sum.cell(row=1, column=1).font = Font(name="Segoe UI", size=15, bold=True, color="1F4E79")
    ws_sum.append([])

    sum_headers = ["ห้อง / แหล่งข้อมูล (Stream)", "วันเปิดตึกจริง", "🟢 สมบูรณ์ 100%", "🟡 มีขาดช่วง", "🟣 ข้อมูลเกิน/ซ้ำ", "🔴 ข้อมูลหาย (Dead)", "คะแนนสุขภาพรวม"]
    ws_sum.append(sum_headers)
    for col_num in range(1, len(sum_headers)+1):
        cell = ws_sum.cell(row=3, column=col_num)
        cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal="center", vertical="center")

    for loc in target_locations:
        st = summary_stats[loc]
        act = st['active_days']
        health_score = (st['healthy'] / act * 100.0) if act > 0 else 0
        
        row_vals = [loc.upper(), f"{act} วัน", f"{st['healthy']} วัน", f"{st['gaps']} วัน", f"{st['excess']} วัน", f"{st['missing']} วัน", f"{health_score:.1f}%"]
        ws_sum.append(row_vals)
        r_idx = ws_sum.max_row
        for c_idx in range(1, len(row_vals)+1):
            c = ws_sum.cell(row=r_idx, column=c_idx)
            c.border = thin_border
            c.font = Font(name="Segoe UI", size=10, bold=(c_idx in [1, 7]))
            c.alignment = Alignment(horizontal="center" if c_idx>1 else "left")
            if c_idx == 7:
                c.fill = PatternFill(start_color="C6EFCE" if health_score>=80 else "FFEB9C" if health_score>=60 else "FFC7CE", fill_type="solid")

    # =========================================================================
    # 📅 SHEET 2: DAILY DIAGNOSTIC MATRIX
    # =========================================================================
    ws_mat = wb.create_sheet(title="📅 Daily Diagnostic Matrix")
    ws_mat.views.sheetView[0].showGridLines = True

    mat_headers = ["วันที่ (Calendar Date)"] + [loc.upper() for loc in target_locations]
    ws_mat.append(mat_headers)
    for col_num in range(1, len(mat_headers)+1):
        cell = ws_mat.cell(row=1, column=col_num)
        cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal="center", vertical="center")

    for r_data, r_status in matrix_rows:
        row_vals = [r_data['Date']] + [r_data[loc.upper()] for loc in target_locations]
        ws_mat.append(row_vals)
        curr_row = ws_mat.max_row

        # แต้มสีตามสถานะที่คำนวณไว้
        date_cell = ws_mat.cell(row=curr_row, column=1)
        date_cell.border, date_cell.alignment, date_cell.font = thin_border, Alignment(horizontal="center"), Font(name="Segoe UI", size=10, bold=True)

        for c_idx, loc in enumerate(target_locations, start=2):
            cell = ws_mat.cell(row=curr_row, column=c_idx)
            status_tag = r_status[loc.upper()]
            cell.fill = fills[status_tag]
            cell.font = fonts[status_tag]
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    print(f"{Colors.CYAN}⏳ [4/4] Auto-adjusting Excel layout dimensions...{Colors.RESET}")
    for ws in [ws_sum, ws_mat]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    wb.save(excel_path)
    print(f"\n{Colors.BOLD}{Colors.GREEN}================================================================={Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.GREEN} 🎉 EXCEL AUDIT REPORT GENERATED SUCCESSFULLY!{Colors.RESET}")
    print(f"    -> {Colors.CYAN}{excel_path}{Colors.RESET}")
    print(f"{Colors.GREEN}=================================================================\n{Colors.RESET}")

if __name__ == "__main__":
    DiagnoseDataHealth()